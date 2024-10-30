# utils/data_parser.py

import os
import matplotlib
import pandas as pd
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

matplotlib.use('Agg')  # Use a non-interactive backend for saving plots
pd.set_option('future.no_silent_downcasting', True)  # Disable future warning for downcasting


def filter_results(results, success=True):
    """
    Filters the test results to include only successful or failed requests.

    :param results: List of dictionaries with request results.
    :param success: Boolean to filter success (True) or failed (False) requests.
    :return: Filtered list of request results.
    """
    return [result for result in results if result["success"] == success]


def extract_response_times(results):
    """
    Extracts response times from the results for analysis.

    :param results: List of dictionaries with request results.
    :return: List of response times (float).
    """
    return [result["response_time"] for result in results if result["response_time"] is not None]


def calculate_percentile(response_times, percentile):
    """
    Calculates a specific percentile for response times (e.g., 90th percentile).

    :param response_times: List of response times.
    :param percentile: Percentile to calculate (e.g., 90 for 90th percentile).
    :return: Calculated percentile value.
    """
    if not response_times:
        return None  # Return None if no response times are available
    sorted_times = sorted(response_times)
    index = int(len(sorted_times) * (percentile / 100))
    return sorted_times[index]


def save_to_csv(results, filepath="test-output/csv-report/test_report.csv"):
    """
    Saves the results to a CSV file for easy viewing and analysis.

    :param results: List of dictionaries with request results.
    :param filepath: File path for the CSV file.
    """
    # Creating a DataFrame from results and explicitly setting data types
    df = pd.DataFrame(results).astype({"response_time": "float", "success": "bool"})

    # Fill NaN values for response time explicitly and confirm as float
    df['response_time'] = df['response_time'].fillna(0).astype(float)
    df['success'] = df['success'].map({True: 'Yes', False: 'No'})  # Convert boolean to Yes/No

    # Adding a summary row with overall metrics
    summary = {
        'Total Requests': len(results),
        'Total Failed Requests': len([r for r in results if r["success"] == "No"]),
        'Average Response Time': df['response_time'].mean(),
        '90th Percentile Response Time': df['response_time'].quantile(0.9)
    }

    summary_df = pd.DataFrame([summary])

    # Saving to CSV
    with open(filepath, mode="w", newline="") as file:
        summary_df.to_csv(file, index=False, header=True)  # Write summary first
        df.to_csv(file, index=False, header=True)  # Then write the main data

    print(f"CSV report saved successfully at: {filepath}")


def generate_visualizations(results):
    """
    Generates visualizations based on the test results and saves them as images.

    :param results: List of dictionaries with request results.
    """
    # Create output directory for images if it doesn't exist
    if not os.path.exists("test-output/visualizations"):
        os.makedirs("test-output/visualizations")

    # Creating DataFrame from results
    df = pd.DataFrame(results)
    df['response_time'] = df['response_time'].fillna(0)  # Replace None with 0 for failed requests

    # Check success column format and convert boolean to Yes/No if needed
    if df['success'].dtype == bool:
        df['success'] = df['success'].map({True: 'Yes', False: 'No'})

    # Separate successful and failed response times for analysis
    successful = df[df['success'] == 'Yes']['response_time']
    failed = df[df['success'] == 'No']['response_time']

    # Plot response times with zero-time highlight for failures
    plt.figure(figsize=(10, 6))
    plt.plot(df.index, df['response_time'], marker='o', linestyle='-', color='b', label='Response Time')
    plt.axhline(y=successful.mean(), color='r', linestyle='--', label='Average Successful Response Time')

    if successful.size > 0:
        percentile_90th = successful.quantile(0.9)  # Calculate 90th percentile
        plt.axhline(y=percentile_90th, color='purple', linestyle=':', label='90th Percentile')

    # Highlight zero response times in red for visibility
    zero_times = df[df['response_time'] == 0]
    if not zero_times.empty:
        plt.scatter(zero_times.index, zero_times['response_time'], color='red', label='Zero Response Time (Failures)')

    plt.title('Response Time Over Requests')
    plt.xlabel('Request Number')
    plt.ylabel('Response Time (seconds)')
    plt.legend()
    plt.grid()
    plt.savefig('test-output/visualizations/response_times.png')
    plt.close()

    # Plot success/failure rates
    success_count = len(successful)
    failure_count = len(failed)

    plt.figure(figsize=(8, 5))
    plt.bar(['Success', 'Failure'], [success_count, failure_count], color=['green', 'red'])
    plt.title('Success vs Failure Rates')
    plt.ylabel('Number of Requests')
    plt.grid(axis='y')
    plt.savefig('test-output/visualizations/success_failure.png')
    plt.close()

    print("Visualizations generated successfully.")


def save_to_excel(results, filepath="test-output/excel-report/test_report.xlsx"):
    """
    Saves the results to an Excel file with color formatting for better visualization.

    :param results: List of dictionaries with request results.
    :param filepath: File path for the Excel file.
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # Create a DataFrame from results
    df = pd.DataFrame(results)

    # Replace boolean values with Yes/No for better readability
    df['success'] = df['success'].replace({True: 'Yes', False: 'No'})

    # Create an Excel writer object
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Write the DataFrame to the Excel file
        df.to_excel(writer, index=False, sheet_name='Test Results')

        # Access the workbook and the active worksheet
        workbook = writer.book
        worksheet = writer.sheets['Test Results']

        # Define colors for success and failure for visual differentiation
        success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        failure_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red

        # Print unique values to debug
        print(df['success'].unique())  # Debugging line to check success values

        # Apply color formatting based on success/failure status
        for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns)):  # Loop through all rows, excluding header
            success_value = row[2].value  # Get value of the 'success' column (4th column)
            if success_value == 'Yes':
                for cell in row:
                    cell.fill = success_fill  # Apply success color
            elif success_value == 'No':
                for cell in row:
                    cell.fill = failure_fill  # Apply failure color

    print(f"Excel report saved successfully at: {filepath}")
